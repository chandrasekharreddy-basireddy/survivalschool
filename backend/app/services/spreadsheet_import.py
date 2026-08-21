"""Shared CSV/XLSX row-parsing for bulk import features (attendance rosters,
external campus timetable sync). Both need the same thing: accept a real
uploaded file, tolerate messy real-world headers (case, whitespace, common
synonyms), and hand back plain dict rows keyed by lowercased header text
instead of forcing every caller to reimplement header matching.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime, time

import openpyxl

from app.core.exceptions import ValidationAppError


def parse_tabular_file(filename: str, content: bytes) -> list[dict[str, str]]:
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        rows = _parse_csv(content)
    elif lower.endswith(".xlsx"):
        rows = _parse_xlsx(content)
    else:
        raise ValidationAppError("Unsupported file type — upload a .csv or .xlsx file.")
    if not rows:
        raise ValidationAppError("The file has no data rows.")
    return rows


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    return [{(k or "").strip().lower(): (v or "").strip() for k, v in raw.items() if k} for raw in reader]


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ValidationAppError("Couldn't read this file as an Excel workbook.") from e
    ws = wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
    out = []
    for raw in rows_iter:
        if raw is None or all(v is None for v in raw):
            continue
        row = {headers[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(raw) if i < len(headers) and headers[i]}
        out.append(row)
    return out


def parse_raw_rows(filename: str, content: bytes) -> list[list[str]]:
    """Like parse_tabular_file, but returns every row as a plain list of
    cell strings with NO header-to-dict transformation — for formats where
    row 0 isn't a header at all (e.g. a "grid" timetable: a title row,
    then day/time rows with one column per room). Blank cells become "".
    """
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        rows = _parse_csv_raw(content)
    elif lower.endswith(".xlsx"):
        rows = _parse_xlsx_raw(content)
    else:
        raise ValidationAppError("Unsupported file type — upload a .csv or .xlsx file.")
    if not rows:
        raise ValidationAppError("The file has no data rows.")
    return rows


def _parse_csv_raw(content: bytes) -> list[list[str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("latin-1")
    return [[(c or "").strip() for c in row] for row in csv.reader(io.StringIO(text))]


def _parse_xlsx_raw(content: bytes) -> list[list[str]]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ValidationAppError("Couldn't read this file as an Excel workbook.") from e
    ws = wb.worksheets[0]
    return _rows_of(ws)


def _rows_of(ws) -> list[list[str]]:
    return [
        ["" if v is None else str(v).strip() for v in row]
        for row in ws.iter_rows(values_only=True)
        if row and any(v is not None for v in row)
    ]


def parse_raw_sheets(filename: str, content: bytes) -> list[list[list[str]]]:
    """Like parse_raw_rows, but for an .xlsx workbook with more than one
    sheet, returns every sheet's rows (e.g. a campus timetable spreadsheet
    that has one grid-format "main schedule" sheet plus several separate
    per-lab schedule sheets, all needing to come in from one upload). A CSV
    has only ever the one implicit "sheet"."""
    lower = (filename or "").lower()
    if lower.endswith(".csv"):
        return [_parse_csv_raw(content)]
    if lower.endswith(".xlsx"):
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as e:
            raise ValidationAppError("Couldn't read this file as an Excel workbook.") from e
        sheets = [_rows_of(ws) for ws in wb.worksheets]
        sheets = [rows for rows in sheets if rows]
        if not sheets:
            raise ValidationAppError("The file has no data rows.")
        return sheets
    raise ValidationAppError("Unsupported file type — upload a .csv or .xlsx file.")


def find_column(row: dict[str, str], candidates: list[str]) -> str | None:
    """First non-empty value among a list of acceptable header-name synonyms."""
    for c in candidates:
        if c in row and row[c]:
            return row[c]
    return None


_DATE_FORMATS = ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y")
_TIME_FORMATS = ("%H:%M:%S", "%H:%M", "%I:%M %p", "%I:%M:%S %p", "%Y-%m-%d %H:%M:%S")


def parse_flexible_date(value: str) -> date:
    value = (value or "").strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date: {value!r}")


def parse_flexible_time(value: str) -> time:
    value = (value or "").strip()
    for fmt in _TIME_FORMATS:
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized time: {value!r}")
